import datetime
import logging

from django.contrib.messages.views import SuccessMessageMixin
from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook

from django.shortcuts import render, get_object_or_404, redirect
from django.utils.decorators import method_decorator
from django.contrib import messages
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy, reverse
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils.encoding import escape_uri_path

from .models import Employee, Department, Vacation, ARCHIVE, RELEVANT, PLANNED
from .forms import VacationForm, EmployeeForm, DepartmentForm, EmployeeUpdateForm
from .openpyxl_config import default_entry_style, remarks_font

logger = logging.getLogger(__name__)

DEFAULT = 0
VIEW = 1
EDIT = 2

MONTHS = [
        'Янв',
        'Фев',
        'Мар',
        'Апр',
        'Май',
        'Июн',
        'Июл',
        'Авг',
        'Сен',
        'Окт',
        'Ноя',
        'Дек',
]


def zip_month_vacation_days(current_department):
    vacation_days_by_month = []
    if len(current_department.vacation_days_by_month) > 0:
        i = 0
        for month in MONTHS:
            vacation_days_by_month.append((month, current_department.vacation_days_by_month[i]))
            i += 1
    else:
        i = 0
        for month in MONTHS:
            vacation_days_by_month.append((month, 0))
            i += 1
    return vacation_days_by_month


class DepartmentCreateView(SuccessMessageMixin, CreateView):
    template_name = 'vacations/add_department.html'
    form_class = DepartmentForm
    success_url = reverse_lazy('vacations:index')
    success_message = 'Отдел успешно добавлен'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get(self, *args, **kwargs):
        user = self.request.user
        if not user.is_staff:
            if not (user.is_department_manager and user.employees_permission_level >= EDIT):  # ++ if not HR
                messages.warning(self.request, 'Недостаточно прав для добавления отдела')
                return redirect('vacations:index')
        return super(DepartmentCreateView, self).get(args, kwargs)

    def post(self, *args, **kwargs):
        user = self.request.user
        if not user.is_staff:
            if not (user.is_department_manager and user.employees_permission_level >= EDIT):  # ++ if not HR
                messages.warning(self.request, 'Недостаточно прав для добавления отдела')
                return redirect('vacations:index')
        return super(DepartmentCreateView, self).post(args, kwargs)

    def get_context_data(self, **kwargs):
        #print('kwargs(views) = ', self.kwargs['employee_id'], type(self.kwargs['employee_id']))
        context = super().get_context_data(**kwargs)

        kwargs['user'] = self.request.user

        context['employees'] = Employee.objects.all()
        context['vacations'] = Vacation.objects.all()
        context['departments'] = Department.objects.all()

        return context

    def get_success_url(self):
        self.object.init_vacation_days()
        return reverse_lazy('vacations:by_department', kwargs={'department_id': self.object.pk})


class DepartmentDeleteView(SuccessMessageMixin, DeleteView):
    template_name = 'vacations/delete_department.html'
    model = Department
    success_url = reverse_lazy('vacations:index')
    success_message = 'Отдел успешно удалён'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get(self, *args, **kwargs):
        user = self.request.user
        if not user.is_staff:
            if not (user.is_department_manager and user.employees_permission_level >= EDIT):
                messages.warning(self.request, 'Недостаточно прав для удаления отдела')
                return redirect('vacations:index')
        return super(DepartmentDeleteView, self).get(args, kwargs)

    def post(self, *args, **kwargs):
        user = self.request.user
        if not user.is_staff:
            if not (user.is_department_manager and user.employees_permission_level >= EDIT):
                messages.warning(self.request, 'Недостаточно прав для удаления отдела')
                return redirect('vacations:index')
        return super(DepartmentDeleteView, self).post(args, kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        kwargs['user'] = self.request.user

        department = Department.objects.get(pk=self.kwargs['pk'])
        context['employees'] = Employee.objects.all()
        context['vacations'] = Vacation.objects.all()
        context['departments'] = Department.objects.all()
        context['current_department'] = department
        context['vacation_days_by_month'] = zip_month_vacation_days(department)

        return context


class EmployeeCreateView(SuccessMessageMixin, CreateView):
    template_name = 'vacations/add_employee.html'
    form_class = EmployeeForm
    #success_url = reverse_lazy('vacations:index')
    success_message = 'Запись успешно добавлена'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get(self, *args, **kwargs):
        user = self.request.user
        department = Department.objects.get(pk=int(self.kwargs['department_id']))
        if not user.is_staff:
            if not (user.employees_permission_level >= EDIT
                    or user.is_department_manager and department == user.bound_employee.department):
                messages.warning(self.request, 'Недостаточно прав для добавления сотрудников')
                return redirect('vacations:index')
        return super(EmployeeCreateView, self).get(args, kwargs)

    def post(self, *args, **kwargs):
        user = self.request.user
        department = Department.objects.get(pk=int(self.kwargs['department_id']))
        if not user.is_staff:
            if not (user.employees_permission_level >= EDIT
                    or user.is_department_manager and department == user.bound_employee.department):
                messages.warning(self.request, 'Недостаточно прав для добавления сотрудников')
                return redirect('vacations:index')
        return super(EmployeeCreateView, self).post(args, kwargs)

    def get_context_data(self, **kwargs):
        #print('kwargs(views) = ', self.kwargs['employee_id'], type(self.kwargs['employee_id']))
        context = super().get_context_data(**kwargs)

        kwargs['user'] = self.request.user

        department = Department.objects.get(pk=int(self.kwargs['department_id']))
        context['employees'] = Employee.objects.all()
        context['vacations'] = Vacation.objects.all()
        context['departments'] = Department.objects.all()
        context['current_department'] = department
        context['vacation_days_by_month'] = zip_month_vacation_days(department)

        return context

    def get_initial(self, *args, **kwargs):
        initial = super(EmployeeCreateView, self).get_initial(**kwargs)
        initial['department'] = get_object_or_404(Department, pk=self.kwargs.get('department_id'))
        return initial

    def get_success_url(self):
        return reverse_lazy('vacations:details', kwargs={'employee_id': self.object.pk})


class EmployeeUpdateView(SuccessMessageMixin, UpdateView):
    template_name = 'vacations/edit_employee.html'
    #success_url = reverse_lazy('vacations:index')
    model = Employee
    form_class = EmployeeUpdateForm
    #fields = ('personnel_number', 'department', 'specialty', 'replaces', 'entry_date', 'max_vacation_days')
    template_name_suffix = '_update_form'
    success_message = 'Запись успешно обновлена'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get(self, *args, **kwargs):
        user = self.request.user
        employee = Employee.objects.get(pk=self.kwargs['pk'])
        if not user.is_staff:
            if not (user.employees_permission_level >= EDIT
                    or user.is_department_manager and user.bound_employee.department == employee.department):
                messages.warning(self.request, 'Недостаточно прав для редактирования данных сотрудников')
                return redirect('vacations:details', kwargs['pk'])
        return super(EmployeeUpdateView, self).get(args, kwargs)

    def post(self, *args, **kwargs):
        user = self.request.user
        employee = Employee.objects.get(pk=self.kwargs['pk'])
        if not user.is_staff:
            if not (user.employees_permission_level >= EDIT
                    or user.is_department_manager and user.bound_employee.department == employee.department):
                messages.warning(self.request, 'Недостаточно прав для редактирования данных сотрудников')
                return redirect('vacations:details', kwargs['pk'])
        return super(EmployeeUpdateView, self).post(args, kwargs)

    def get_initial(self):
        base_initial = super().get_initial()
        return base_initial

    def get_context_data(self, **kwargs):
        #print('kwargs(views) = ', self.kwargs['employee_id'], type(self.kwargs['employee_id']))
        context = super().get_context_data(**kwargs)

        kwargs['user'] = self.request.user

        employee = Employee.objects.get(pk=int(self.kwargs['pk']))
        context['employee'] = employee
        context['employees'] = Employee.objects.all()
        context['vacations'] = Vacation.objects.all()
        context['departments'] = Department.objects.all()
        context['current_department'] = employee.department
        context['vacation_days_by_month'] = zip_month_vacation_days(employee.department)

        return context

    def get_success_url(self):
        return reverse_lazy('vacations:details', kwargs={'employee_id': self.object.pk})


class EmployeeDeleteView(SuccessMessageMixin, DeleteView):
    template_name = 'vacations/delete_employee.html'
    model = Employee
    success_url = reverse_lazy('vacations:index')
    success_message = 'Запись успешно удалена'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get(self, *args, **kwargs):
        user = self.request.user
        employee = Employee.objects.get(pk=self.kwargs['pk'])
        if not user.is_staff:
            if not (user.employees_permission_level >= EDIT
                    or user.is_department_manager and user.bound_employee.department.pk == employee.department.pk):  # ++ if not HR
                messages.warning(self.request, 'Недостаточно прав для удаления сотрудников')
                return redirect('vacations:details', kwargs['employee_id'])
        return super(EmployeeDeleteView, self).get(args, kwargs)

    def post(self, *args, **kwargs):
        user = self.request.user
        employee = Employee.objects.get(pk=self.kwargs['pk'])
        if not user.is_staff:
            if not (user.employees_permission_level >= EDIT
                    or user.is_department_manager and user.bound_employee.department.pk == employee.department.pk):  # ++ if not HR
                messages.warning(self.request, 'Недостаточно прав для удаления сотрудников')
                return redirect('vacations:details', kwargs['employee_id'])
        return super(EmployeeDeleteView, self).post(args, kwargs)

    def get_context_data(self, **kwargs):
        #print('kwargs(views) = ', self.kwargs['employee_id'], type(self.kwargs['employee_id']))
        context = super().get_context_data(**kwargs)

        kwargs['user'] = self.request.user

        employee = Employee.objects.get(pk=int(self.kwargs['pk']))
        context['employee'] = employee
        context['employees'] = Employee.objects.all()
        context['vacations'] = Vacation.objects.all()
        context['departments'] = Department.objects.all()
        context['current_department'] = employee.department
        context['vacation_days_by_month'] = zip_month_vacation_days(employee.department)

        return context


class VacationCreateView(SuccessMessageMixin, CreateView):
    template_name = 'vacations/details.html'
    form_class = VacationForm
    success_message = 'Запись успешно добавлена'
    #employee_id = '0'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get(self, *args, **kwargs):
        user = self.request.user
        employee = Employee.objects.get(pk=self.kwargs['employee_id'])
        if not user.is_staff:
            if not (user.employees_permission_level >= VIEW
                    or user.bound_employee.pk == self.kwargs['employee_id']
                    or user.is_department_manager and user.bound_employee.department.pk == employee.department.pk):
                messages.warning(self.request, 'Недостаточно прав для доступа к странице')
                return redirect('vacations:by_department', user.bound_employee.department.pk)
        return super(VacationCreateView, self).get(args, kwargs)

    def post(self, *args, **kwargs):
        user = self.request.user
        employee = Employee.objects.get(pk=self.kwargs['employee_id'])
        if not user.is_staff:
            if not (user.is_department_manager and user.bound_employee.department.pk == employee.department.pk):
                messages.warning(self.request, 'Недостаточно прав для добавления отпусков')
                return redirect('vacations:by_department', user.bound_employee.department.pk)
        return super(VacationCreateView, self).post(args, kwargs)

    def get_success_url(self):
        employee_id = self.kwargs['employee_id']
        #print('get_success_url employee_id = ', employee_id)
        return reverse_lazy('vacations:details', kwargs={'employee_id': employee_id})

    def get_initial(self, *args, **kwargs):
        initial = super(VacationCreateView, self).get_initial(**kwargs)
        initial['employee'] = get_object_or_404(Employee, pk=self.kwargs.get('employee_id'))
        initial['relevance'] = ARCHIVE
        return initial

    def get_context_data(self, **kwargs):
        #print('kwargs(views) = ', self.kwargs['employee_id'], type(self.kwargs['employee_id']))
        context = super().get_context_data(**kwargs)

        kwargs['user'] = self.request.user

        employee = Employee.objects.get(pk=self.kwargs['employee_id'])
        context['employee'] = employee
        context['employees'] = Employee.objects.all()
        context['vacations'] = Vacation.objects.filter(employee=employee)
        context['departments'] = Department.objects.all()
        context['current_department'] = employee.department
        context['vacation_days_by_month'] = zip_month_vacation_days(employee.department)

        return context


class VacationDeleteView(SuccessMessageMixin, DeleteView):
    template_name = 'vacations/delete_vacation.html'
    model = Vacation
    #success_url = reverse_lazy('vacations:index')
    success_message = 'Запись успешно удалена'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get(self, *args, **kwargs):
        user = self.request.user
        vacation = Vacation.objects.get(pk=self.kwargs['pk'])
        if not user.is_staff:
            if not (user.employees_permission_level >= EDIT or user.is_department_manager and user.bound_employee.department.pk == vacation.employee.department.pk):
                messages.warning(self.request, 'Недостаточно прав для удаления отпусков')
                return redirect('vacations:details', kwargs['pk'])
        return super(VacationDeleteView, self).get(args, kwargs)

    def post(self, *args, **kwargs):
        user = self.request.user
        vacation = Vacation.objects.get(pk=self.kwargs['pk'])
        if not user.is_staff:
            if not (user.employees_permission_level >= EDIT or user.is_department_manager and user.bound_employee.department.pk == vacation.employee.department.pk):
                messages.warning(self.request, 'Недостаточно прав для удаления отпусков')
                return redirect('vacations:details', kwargs['pk'])
        return super(VacationDeleteView, self).post(args, kwargs)

    def form_valid(self, *args, **kwargs):
        logger.debug('DeleteView.form_valid()')
        self.object = self.get_object()
        self.object.change_values(reverse=True)
        return super(VacationDeleteView, self).form_valid(*args, **kwargs)

    def get_success_url(self):
        employee_id = Vacation.objects.get(pk=self.kwargs['pk']).employee.pk
        #print('get_success_url employee_id = ', employee_id)
        return reverse_lazy('vacations:details', kwargs={'employee_id': employee_id})

    def get_context_data(self, **kwargs):
        #print('kwargs(views) = ', self.kwargs['employee_id'], type(self.kwargs['employee_id']))
        context = super().get_context_data(**kwargs)

        kwargs['user'] = self.request.user

        vacation = Vacation.objects.get(pk=int(self.kwargs['pk']))
        employee = vacation.employee
        context['employee'] = employee
        context['employees'] = Employee.objects.all()
        context['vacations'] = Vacation.objects.all()
        context['departments'] = Department.objects.all()
        context['current_department'] = employee.department
        context['vacation_days_by_month'] = zip_month_vacation_days(employee.department)

        return context


@login_required
def by_department(request, department_id):
    employees = Employee.objects.filter(department=department_id)
    vacations = Vacation.objects.all()
    departments = Department.objects.all()
    current_department = Department.objects.get(pk=department_id)

    vacation_days_by_month = zip_month_vacation_days(current_department)

    user = request.user

    logger.debug(int(datetime.date.today().year - 1))

    #current_department.test_vacation_days()

    if not user.is_staff:
        if not (user.employees_permission_level >= VIEW
                or user.bound_employee.department.pk == department_id):
            #messages.warning(request, 'Недостаточно прав для доступа к странице')
            return redirect('vacations:details', user.bound_employee.pk)

    context = {'employees': employees, 'departments': departments, 'vacations': vacations,
               'current_department': current_department, 'vacation_days_by_month': vacation_days_by_month}
    return render(request, 'vacations/by_department.html', context)


@login_required
def recalculate_department(request, department_id):
    employees = Employee.objects.filter(department=department_id)
    current_department = Department.objects.get(pk=department_id)

    user = request.user

    if not user.is_staff:
        if not (user.employees_permission_level >= VIEW
                or (user.bound_employee.department.pk == department_id and user.is_department_manager)):
            #messages.warning(request, 'Недостаточно прав для доступа к странице')
            return redirect('vacations:details', user.bound_employee.pk)

    current_department.init_vacation_days()

    for employee in employees:
        employee.reset_rating()
        employee.reset_vacation_days()
        #logger.debug(employee.rating)
        employee.save()

        vacations = Vacation.objects.filter(employee=employee.pk)

        for vacation in vacations:
            vacation.adjust_relevance()
            #if vacation.relevance == PLANNED:
            #    current_department.change_vacation_days(vacation.start, vacation.end)

    return redirect('vacations:by_department', department_id)


def generate_t7_form(department_id=None, year=None):
    if department_id is not None:
        departments = [Department.objects.get(pk=department_id)]
    else:
        departments = Department.objects.all()

    wb = load_workbook('export_forms/t7_template.xlsx')
    # ...
    ws = wb.active

    wb.add_named_style(default_entry_style)

    entries_count = 0
    current_row = 20

    for department in departments:
        employees = Employee.objects.filter(department=department.pk)
        for employee in employees:
            if year is None:
                employee_vacations = Vacation.objects.filter(employee=employee.pk, relevance=PLANNED)
            else:
                employee_vacations = Vacation.objects.filter(employee=employee.pk, start__year=year)
            employee_vacations_count = 0
            current_row = 20 + entries_count
            for vacation in employee_vacations:
                delta = vacation.end - vacation.start
                vacation_calendar_days = delta.days + 1

                current_row = 20 + entries_count

                ws.insert_rows(current_row)

                if type(ws.cell(row=current_row, column=1)).__name__ == 'MergedCell':
                    logger.debug('merged cell detected')
                    ws.unmerge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
                ws.cell(row=current_row, column=1, value=' '+str(department))

                if employee.specialty is not None:
                    specialty_str = employee.specialty
                else:
                    specialty_str = ''
                ws.cell(row=current_row, column=2, value=' '+specialty_str)
                ws.cell(row=current_row, column=3, value=' '+str(employee))
                ws.cell(row=current_row, column=4, value=' '+str(employee.personnel_number))
                ws.cell(row=current_row, column=5, value=' '+str(vacation_calendar_days))
                ws.cell(row=current_row, column=7, value=' '+str(vacation.start))
                ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=6)
                ws.cell(row=current_row, column=8, value='')
                ws.cell(row=current_row, column=9, value='')
                ws.cell(row=current_row, column=12, value='')
                ws.cell(row=current_row, column=15, value='')

                ws.merge_cells(start_row=current_row, start_column=15, end_row=current_row, end_column=17)
                ws.merge_cells(start_row=current_row, start_column=9, end_row=current_row, end_column=11)
                ws.merge_cells(start_row=current_row, start_column=12, end_row=current_row, end_column=14)

                ws.row_dimensions[current_row].height = 30

                entries_count += 1
                employee_vacations_count += 1

            if employee_vacations_count > 1:
                ws.merge_cells(start_row=current_row - employee_vacations_count + 1, start_column=1,
                               end_row=current_row, end_column=1)
                ws.merge_cells(start_row=current_row - employee_vacations_count + 1, start_column=2,
                               end_row=current_row, end_column=2)
                ws.merge_cells(start_row=current_row - employee_vacations_count + 1, start_column=3,
                               end_row=current_row, end_column=3)
                ws.merge_cells(start_row=current_row - employee_vacations_count + 1, start_column=4,
                               end_row=current_row, end_column=4)

    if entries_count % 2 == 1:
        _init = 2
    else:
        _init = 1

    ws.row_dimensions[current_row + _init].height = 35
    ws.row_dimensions[current_row + _init + 1].height = 12
    ws.row_dimensions[current_row + _init + 2].height = 35
    ws.row_dimensions[current_row + _init + 3].height = 12
    ws.row_dimensions[current_row + _init + 4].height = 35
    ws.row_dimensions[current_row + _init + 5].height = 12

    for row in ws.iter_rows(min_row=20, max_col=17, max_row=current_row-2+_init):
        for cell in row:
            cell.style = 'default_entry_style'
            if cell.column == 15:
                cell.font = remarks_font

    return wb


@login_required
def export_t7_department(request, department_id, year):
    user = request.user

    if not user.is_staff:
        if not (user.employees_permission_level >= VIEW
                or (user.bound_employee.department.pk == department_id and user.is_department_manager)):
            #messages.warning(request, 'Недостаточно прав для доступа к странице')
            return redirect('vacations:details', user.bound_employee.pk)

    if year == datetime.date.today().year + 1:
        wb = generate_t7_form(department_id=department_id)
    else:
        wb = generate_t7_form(department_id=department_id, year=year)

    department_title = Department.objects.get(pk=department_id).title
    filename = f'Форма Т-7 (График отпусков) {department_title} {year}.xlsx'

    response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/ms-excel')
    response['Content-Disposition'] = "attachment; filename=" + escape_uri_path(filename)
    return response


@login_required
def export_t7_all(request, year):
    user = request.user

    if not user.is_staff:
        if user.employees_permission_level < VIEW:
            return redirect('vacations:details', user.bound_employee.pk)

    if year == datetime.date.today().year + 1:
        wb = generate_t7_form()
    else:
        wb = generate_t7_form(year=year)

    filename = f'Форма Т-7 (График отпусков) {year}.xlsx'

    response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/ms-excel')
    response['Content-Disposition'] = "attachment; filename=" + escape_uri_path(filename)
    return response


'''
@login_required
def details(request, employee_id):
    current_employee = Employee.objects.get(pk=employee_id)
    employees = Employee.objects.all()
    vacations = Vacation.objects.all()
    departments = Department.objects.all()
    context = {'employees': employees, 'vacations': vacations,
               'departments': departments, 'employee': current_employee}
    return render(request, 'vacations/details.html', context)
'''


@login_required
def index(request):
    employees = Employee.objects.all()
    vacations = Vacation.objects.all()
    departments = Department.objects.all()

    user = request.user

    if not (user.employees_permission_level >= VIEW or user.is_staff):  # ++ if not HR
        #messages.warning(request, 'Недостаточно прав для доступа к странице')
        return redirect('vacations:by_department', user.bound_employee.department.pk)

    context = {'employees': employees, 'vacations': vacations, 'departments': departments}
    return render(request, 'vacations/index.html', context)


def success(request):
    return render(request, 'vacations/success.html')


def bootstrap_test(request):
    return render(request, 'vacations/test1.html')
